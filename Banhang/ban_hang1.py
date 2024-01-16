import sys
from PyQt5.QtWidgets import QApplication, QMainWindow, QPushButton, QVBoxLayout, QWidget, QLineEdit, QCompleter, \
    QTableWidgetItem, QTableWidget, QMessageBox,QCompleter
from PyQt5.QtCore import QStringListModel, Qt,QUrl
from PyQt5.QtGui import QDesktopServices,QPixmap
import pandas as pd
from openpyxl import load_workbook
from ban import Ui_MainWindow
from num2words import num2words
import locale

class BanHangApp(QMainWindow, Ui_MainWindow):
    def __init__(self):
        super().__init__()
        self.setupUi(self)
        self.pushButton_co_ma.clicked.connect(self.open_excel_file1)
        self.pushButton_khong_ma.clicked.connect(self.open_excel_file2)
        self.pushButton_thoat.clicked.connect(self.close)
        self.pushButton_xoa_dong.clicked.connect(self.delete_selected_row)
        self.pushButton_xoa.clicked.connect(self.delete_row)
        self.pushButton_quet.clicked.connect(self.check_and_display)
        self.pushButton_nhap.clicked.connect(self.khong_ma)
        self.pushButton_thanh_toan.clicked.connect(self.on_thanhtoan_clicked)
        self.lineEdit_ma_vach.setFocus()

        self.lineEdit_ma_vach.returnPressed.connect(self.check_and_display)
        pd.set_option('display.float_format', lambda x: '%.0f' % x)
        self.lineEdit_khong_ma.textChanged.connect(self.auto_complete)
        border_style = "QLabel { border: 1px solid black; padding: 5px; }"
        self.label_tien_so.setStyleSheet(border_style)
        self.label_tien_chu.setStyleSheet(border_style)
        # Dòng hiện tại đang được quét
        self.current_row = 1
        self.set_image()

    def set_image(self):
        # Load the image from file (replace 'path/to/your/image.jpg' with the actual file path)
        image_path = 'be.jpg'
        pixmap = QPixmap(image_path)

        # Set the pixmap to the label
        self.label_logo.setPixmap(pixmap)

    def open_excel_file1(self):
        excel_file_path = r'D:\Banhang\Excel\banggia.xlsx'
        QDesktopServices.openUrl(QUrl.fromLocalFile(excel_file_path))

    def open_excel_file2(self):
        excel_file_path = r'D:\Banhang\Excel\khong_ma.xlsx'
        QDesktopServices.openUrl(QUrl.fromLocalFile(excel_file_path))

    def khong_ma(self):
        wb1 = load_workbook('D:\\Banhang\\Excel\\khong_ma.xlsx')
        ws1 = wb1['Sheet1']
        wtsp1 = ws1['a']
        wmsp1 = ws1['b']
        wgsp1 = ws1['c']
        entered_value1 = self.lineEdit_khong_ma.text()
        try:
            found = False
            for i in range(1, len(wtsp1)):
                msp1 = str(wmsp1[i].value)
                tsp1 = str(wtsp1[i].value)
                gsp1 = str(wgsp1[i].value)
                sl = "1"
                if str(tsp1) == str(entered_value1):
                    row_position = self.tableWidget.rowCount()
                    self.tableWidget.insertRow(row_position)
                    self.lineEdit_khong_ma.clear()
                    self.tableWidget.setItem(row_position, 0, QTableWidgetItem(tsp1))
                    self.tableWidget.setItem(row_position, 1, QTableWidgetItem(msp1))
                    self.tableWidget.setItem(row_position, 2, QTableWidgetItem(gsp1))
                    self.tableWidget.setItem(row_position, 3, QTableWidgetItem(sl))

                    # Optional: Auto-scroll to the newly added row
                    self.tableWidget.scrollToItem(self.tableWidget.item(row_position, 0))
                    found = True
                    break

            if not found:
                # Hiển thị cửa sổ thông báo nếu Mã hàng không tồn tại
                QMessageBox.warning(self, "Thông báo", "Mã hàng không tồn tại sếp ơi !", QMessageBox.Ok)
                self.lineEdit_khong_ma.clear()

        except Exception as e:
            print(e)

    def auto_complete(self, text):
        try:
            wb1 = load_workbook('D:\\Banhang\\Excel\\khong_ma.xlsx')
            ws1 = wb1['Sheet1']
            wtsp1 = ws1['a']
            wmsp1 = ws1['b']
            wgsp1 = ws1['c']
            tsp_values = [str(cell.value) for cell in wtsp1 if cell.value is not None and str(cell.value).strip()]
            # Sử dụng QCompleter để tạo gợi ý cho line_edit_khong_ma
            completer = QCompleter(tsp_values, self)
            completer.setCaseSensitivity(Qt.CaseInsensitive)
            # Gán QCompleter cho line_edit_khong_ma
            self.lineEdit_khong_ma.setCompleter(completer)
        except Exception as e:
            print("Lỗi khi đọc dữ liệu:", e)

    def check_and_display(self):
        wb = load_workbook('D:\\Banhang\\Excel\\banggia.xlsx')
        ws = wb['Sheet1']
        wtsp = ws['a']
        wmsp = ws['b']
        wgsp = ws['c']
        entered_value = self.lineEdit_ma_vach.text()

        try:
            found = False
            for i in range(1, len(wmsp)):
                msp = str(wmsp[i].value)
                tsp = str(wtsp[i].value)
                gsp = str(wgsp[i].value)
                sl = "1"
                if str(msp) == str(entered_value):
                    row_position = self.tableWidget.rowCount()
                    self.tableWidget.insertRow(row_position)
                   self.lineEdit_ma_vach.clear()
                    # Add data to the cells in the new row
                    self.tableWidget.setItem(row_position, 0, QTableWidgetItem(tsp))
                    self.tableWidget.setItem(row_position, 1, QTableWidgetItem(msp))
                    self.tableWidget.setItem(row_position, 2, QTableWidgetItem(gsp))
                    self.tableWidget.setItem(row_position, 3, QTableWidgetItem(sl))
                    # Optional: Auto-scroll to the newly added row
                    self.tableWidget.scrollToItem(self.tableWidget.item(row_position, 0))
                    found = True
                    break

            if not found:
                # Hiển thị cửa sổ thông báo nếu Mã hàng không tồn tại
                QMessageBox.warning(self, "Thông báo", "Mã hàng không tồn tại sếp ơi !", QMessageBox.Ok)
                self.lineEdit_ma_vach.clear()

        except Exception as e:
            print(e)

    def delete_selected_row(self):
        try:
            selected_row = self.tableWidget.currentRow()
            if selected_row != -1 and selected_row < self.tableWidget.rowCount():
                # Xóa dòng được chọn
                self.tableWidget.removeRow(selected_row)
                self.current_row -= 1
            else:
                QMessageBox.warning(self, "Thông báo", "Sếp vui lòng chọn dòng để xóa!", QMessageBox.Ok)
        except Exception as e:
            print(e)

    def delete_row(self):
        try:
            # Đặt số lượng dòng của QTableWidget về 0 để xóa toàn bộ dòng
            self.tableWidget.setRowCount(0)
            self.current_row = 1  # Đặt lại chỉ số dòng hiện tại nếu cần thiết
            self.label_tien_so.setText("")
            self.label_tien_chu.setText("")
        except Exception as e:
            print(e)

    def calculate_total(self):
        total = 0
        for row in range(self.tableWidget.rowCount()):
            price = float(self.tableWidget.item(row, 2).text())
            quantity = int(self.tableWidget.item(row, 3).text())
            total += price * quantity
        return total

    def on_thanhtoan_clicked(self):
        try:
            total_value = self.calculate_total()
            total_integer = int(total_value)
            # Sử dụng num2words để chuyển đổi số sang chữ tiếng Việt
            total_words = num2words(total_integer, lang='vi')
            # Sử dụng locale để định dạng số với dấu phẩy ngăn cách
            locale.setlocale(locale.LC_ALL, 'vi_VN.utf8')
            total_formatted = locale.format("%d", total_integer, grouping=True)

            # Hiển thị giá trị chữ và số lên label
            self.label_tien_so.setText(total_formatted)
            self.label_tien_chu.setText(total_words)  # Tự viết hoa chữ cái đầu tiên
        except Exception as e:
            print(e)

if __name__ == "__main__":
    app = QApplication(sys.argv)
    mainWin = BanHangApp()
    mainWin.show()
    sys.exit(app.exec_())

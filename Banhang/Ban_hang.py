import tkinter as tk
from tkinter import*
from tkinter import ttk
import xlrd
from openpyxl import Workbook,load_workbook
import os
import time
from tkinter.font import Font
from PIL import ImageTk, Image
from itertools import islice
import datetime
from num2words import num2words


class AutocompleteCombobox(ttk.Combobox):

        def set_completion_list(self, completion_list):
                """Use our completion list as our drop down selection menu, arrows move through menu."""
                self._completion_list = sorted(completion_list, key=str.lower) # Work with a sorted list
                self._hits = []
                self._hit_index = 0
                self.position = 0
                self.bind('<KeyRelease>', self.handle_keyrelease)
                self['values'] = self._completion_list  # Setup our popup menu

        def autocomplete(self, delta=0):
                """autocomplete the Combobox, delta may be 0/1/-1 to cycle through possible hits"""
                if delta: # need to delete selection otherwise we would fix the current position
                        self.delete(self.position, tk.END)
                else: # set position to end so selection starts where textentry ended
                        self.position = len(self.get())
                # collect hits
                _hits = []
                for element in self._completion_list:
                        if element.lower().startswith(self.get().lower()): # Match case insensitively
                                _hits.append(element)
                # if we have a new hit list, keep this in mind
                if _hits != self._hits:
                        self._hit_index = 0
                        self._hits=_hits
                # only allow cycling if we are in a known hit list
                if _hits == self._hits and self._hits:
                        self._hit_index = (self._hit_index + delta) % len(self._hits)
                # now finally perform the auto completion
                if self._hits:
                        self.delete(0,tk.END)
                        self.insert(0,self._hits[self._hit_index])
                        self.select_range(self.position,tk.END)

        def handle_keyrelease(self, event):
                """event handler for the keyrelease event on this widget"""
                if event.keysym == "BackSpace":
                        self.delete(self.index(tk.INSERT), tk.END)
                        self.position = self.index(tk.END)
                if event.keysym == "Left":
                        if self.position < self.index(tk.END): # delete the selection
                                self.delete(self.position, tk.END)
                        else:
                                self.position = self.position-1 # delete one character
                                self.delete(self.position, tk.END)
                if event.keysym == "Right":
                        self.position = self.index(tk.END) # go to end (no selection)
                if len(event.keysym) == 1:
                        self.autocomplete()

root = tk.Tk(className='AutocompleteCombobox')

root.title("Bán hàng bố Sinh")


def update_time():
	# Get the current date and time
	lb5=tk.Label(root,text="Ngày giờ",font=("Arial",15),bg="lightskyblue")
	lb5.place(x=70,y=700)
	entry5 = Entry(root, width=18, font=("Calibri", 18),bg="lightskyblue")
	entry5.place(x=1,y=750)

	date = datetime.datetime.now()
	# Format the date and time
	hom_nay = date.strftime("%d-%m-%Y ___ %H:%M")
	# Update the Entry widget with the new date and time
	entry5.delete(0, END)
	entry5.insert(END, hom_nay)
	# Call the function again after 60 seconds
	entry5.after(30000, update_time)

# Call the function to start updating the date and time
update_time()

w = 1500 # width for the Tk root
h = 800 # height for the Tk root

# get screen width and height
ws = root.winfo_screenwidth() # width of the screen
hs = root.winfo_screenheight() # height of the screen

# calculate x and y coordinates for the Tk root window
x = (ws/2) - (w/2)
y = (hs/2) - (h/2)

# set the dimensions of the screen
# and where it is placed
photo = ImageTk.PhotoImage(Image.open("D:\\Banhang\\th.jpg"))

# label = tk.Label(root, image=photo)
# label.place(x=10,y=10)

canvas = tk.Canvas(root, width=150, height=150)
canvas.place(x=10,y=10)
canvas.create_image(65, 65, image=photo)

root.geometry('%dx%d+%d+%d' % (w, h, x, y))

wb = load_workbook('D:\\Banhang\\Excel\\banggia.xlsx')
ws = wb['Sheet1']
wtsp = ws['a']
wmsp = ws['b']
wgsp = ws['c']

wb1 = load_workbook('D:\\Banhang\\Excel\\khong_ma.xlsx')
ws1 = wb1['Sheet1']
wtsp1 = ws1['a']
wmsp1 = ws1['b']
wgsp1 = ws1['c']

cbb = AutocompleteCombobox(root,width=25,height=16,font=("Arial",13))
cbb.place(x=1150,y=100)

root.option_add("*TCombobox*Listbox.font", "Calibri 14")
values = []

for cell in islice(wtsp1, 1):
	values.append(cell.value)
cbb.set_completion_list(values)

cbb.focus_set()

label = tk.Label(root,text = "Tạp Hóa Sinh Niềm " ,font = ("Times",20),bg="lightskyblue")
label.place(x = 650, y = 30)

label1 = tk.Label(root, text="Mã vạch đang quét  :",font= ("Arial", 16),bg="lightskyblue")
label1.place(x=500,y=100)

label2 = tk.Label(root,text = "product by TuocNguyen ",font = ("Arial",14),fg="blue",bg="lightskyblue" )
label2.place(x=1200,y=750)

text = tk.StringVar()
entry =Entry(root, textvariable=text,font= ("Arial", 16),fg="blue")
entry.place(x=750, y=100)
input_value = text.get()
entry.focus()

def get_entry_value():
	input_value = text.get()
	quet_mv()

button = tk.Button(root, text="Quét", command=get_entry_value,font='Arrial',width=5,bg="green",fg="white")
button.place(x=1050, y=97)

continue_scan = True
# Tạo một khung để chứa lưới và thanh cuộn
frame = tk.Frame(root)
frame.place(x=350, y=200)

tree = ttk.Treeview(frame, height=15, show="tree headings", style="Custom.Treeview")
# Tạo một đối tượng Style
style = ttk.Style()

# Thay đổi màu chữ của toàn bộ lưới thành xanh lá cây
style.configure("Custom.Treeview", foreground="blue")
# Tạo một đối tượng Font với tên font là Arial, kích thước font là 16, kiểu font là in đậm
font = Font(family="Arial", size=13)
font1 = Font(family="Arial", size=16, weight="bold")
# Tạo một nhãn có tên là "bigrow" với thuộc tính font là đối tượng font đã tạo
tree.tag_configure("bigrow", font=font)
tree.tag_configure("bigrow1", font=font1)
style.configure("Custom.Treeview.Heading", font=font1)

tree["columns"] = ("one", "two", "three")
tree.column("#0", width=0, minwidth=5, stretch=tk.NO)
tree.column("one", width=300, minwidth=7, stretch=tk.NO)
tree.column("two", width=230,minwidth=7,stretch=tk.NO)
tree.column("three", width=230,minwidth=7, stretch=tk.NO)
tree.column("#0", anchor=tk.CENTER)
tree.column("one", anchor=tk.CENTER)
tree.column("two", anchor=tk.CENTER)
tree.column("three", anchor=tk.CENTER)

# Tạo các tiêu đề cho các cột
tree.heading("#0", text="", anchor=tk.W)
tree.heading("one", text="Tên hàng", anchor=tk.N)
tree.heading("two", text="Mã hàng", anchor=tk.N)
tree.heading("three", text="Giá bán", anchor=tk.N)

tree.grid(row=0, column=0)

# Tạo một thanh cuộn ngang bằng ttk.Scrollbar
x_scrollbar = ttk.Scrollbar(frame, orient="horizontal", command=tree.xview)

# Đặt thanh cuộn ngang vào khung
x_scrollbar.grid(row=1, column=0, sticky="ew")

# Kết nối thanh cuộn ngang với lưới
tree.configure(xscrollcommand=x_scrollbar.set)

# Tạo một thanh cuộn dọc bằng ttk.Scrollbar
y_scrollbar = ttk.Scrollbar(frame, orient="vertical", command=tree.yview)

# Đặt thanh cuộn dọc vào khung
y_scrollbar.grid(row=0, column=1, sticky="ns")

# Kết nối thanh cuộn dọc với lưới
tree.configure(yscrollcommand=y_scrollbar.set)

def enable_entry():
	# Đặt thuộc tính state thành NORMAL để cho phép nhập liệu
	entry.config(state="normal")

	entry.focus()

# Tạo một hàm để vô hiệu hóa widget Entry
def disable_entry():
	# Đặt thuộc tính state thành DISABLED để không cho nhập liệu
	entry.config(state="disabled")
	# Gọi hàm quet_mv() sau khi quét xong
	quet_mv()

def open_excel3():
	# Đọc file excel có sẵn
	os.startfile('D:\\Banhang\\Excel\\banggia.xlsx')

def open_excel4():
	# Đọc file excel có sẵn
	os.startfile('D:\\Banhang\\Excel\\khong_ma.xlsx')

current_row = 0

def quet_mv():
	global current_row

	global continue_scan # Khai báo biến toàn cục continue_scan

	input_value = text.get()
	for i in range(len(wmsp)):
		msp = wmsp[i].value
		tsp = wtsp[i].value
		gsp = wgsp[i].value

		if str(msp) == str(input_value):

			time.sleep(1)

			entry.after_idle(enable_entry)

			text.set("")

			entry.focus()

			tree.insert(parent="", index="end", iid=current_row, text="", values=(tsp, input_value, gsp),tags=("bigrow",))

			current_row += 1
			i+=1
		else:
			enable_entry()
			text.set("")
			entry.focus()

def tong():

	global current_row
	input_value = cbb.get()

	for i in range(len(wmsp1)):
		msp1 = wmsp1[i].value
		tsp1 = wtsp1[i].value
		gsp1 = wgsp1[i].value

		if str(tsp1) == str(input_value):

			tree.insert(parent="", index="end", iid=current_row, text="", values=(input_value,msp1, gsp1),tags=("bigrow",))

			current_row += 1
			break
		cbb.set("")

cbb['values'] = values
button1= tk.Button(root,text="Nhập",command=tong,font='Arrial',width=5,fg="white",bg="green")
button1.place(x=1430,y=97)

def sum_values():
	global total_row # Khai báo biến toàn cục total_row
	total = 0
	for row in tree.get_children():
		value = tree.item(row, "values")[2]
		# Chuyển đổi giá trị từ chuỗi sang số nguyên bằng cách loại bỏ dấu phẩy
		value = int(value.replace(",", ""))
		total += (value)
		words_vi = num2words(total, lang='vi')

		lb1=Label(root,text="Thanh toán ",font=("Arial bold",14),width=9,bg="lightskyblue",fg="blue")
		lb1.place(x=1250,y=400)
		lb=Label(root,text=total,font=("Arial bold",14),width=8,bg="white",fg="blue")
		lb.place(x=1250,y=450)
		lb=Label(root,text=words_vi,font=("Arial bold",14),width=28,bg="white",fg="blue")
		lb.place(x=1150,y=500)


	total_row = None
def delete_row():

	lb1=Label(root,text="",font=("Arial bold",14),width=9,bg="lightskyblue",fg="blue")
	lb1.place(x=1250,y=400)
	lb=Label(root,text="",font=("Arial bold",14),width=8,bg="lightskyblue",fg="blue")
	lb.place(x=1250,y=450)
	lb=Label(root,text="",font=("Arial bold",14),width=28,bg="lightskyblue",fg="blue")
	lb.place(x=1150,y=500)
	# Lấy tất cả các hàng trong lưới bằng phương thức get_children()
	all_items = tree.get_children()
	# Xóa tất cả các hàng khỏi lưới bằng phương thức delete()
	for item in all_items:
		tree.delete(item)
	entry.focus()

def delete():
	# Lấy id của hàng đã chọn hoặc tất cả các hàng trong lưới
	selected_items = tree.selection()
	# Nếu có hàng nào được chọn, xóa nó khỏi lưới
	if selected_items:
		for item in selected_items:
		# Tìm vị trí của dòng đã chọn trong lưới
			row_index = tree.index(item)
		# Xóa dòng tương ứng trong sheet
			ws.delete_rows(row_index + 2)
		# Xóa dòng khỏi lưới
			tree.delete(item)
	else:

		new_win1 = tk.Toplevel(root)
		new_win1.geometry("200x100")
		new_win1.title("Xóa dữ liệu")
		new_win1.transient(root)


		label_1 = tk.Label(new_win1, text="Bạn chưa chọn hàng nào để xóa")
		label_1.place(x=5,y=15)
		def ket_thuc1() :

			new_win1.destroy()

		button_insert = tk.Button(new_win1, text="OK",width=10, command=ket_thuc1)
		button_insert.place(x=64,y=45)

button = tk.Button(root, text="Xóa dòng ", command=delete,font='Arrial',width=9)
button.place(x=1200, y=200)
def thoat():
	root.destroy()
button = tk.Button(root, text="Thoát", command=thoat,font='Arrial',width=9)
button.place(x=950, y=650)
button = tk.Button(root, text="Hàng có mã", command=open_excel3,font='Arrial',width=9)
button.place(x=550, y = 590)
button = tk.Button(root, text="Hàng không mã", command=open_excel4,font='Arrial',width=12)
button.place(x=550, y = 650)

button = tk.Button(root, text="Xóa", command=delete_row,font='Arrial',width=9)
button.place(x=950, y=590)
button = tk.Button(root, text="Thanh toán", command=sum_values,font='Arrial',width=9,bg="green",fg="white")
button.place(x=750, y=590)
def validate_input(P):
	# Nếu P là chuỗi rỗng hoặc là số nguyên, trả về True
	if P == "" or P.isdigit():
		return True
	# Nếu không, trả về False
	else:
		return False

# Đăng ký một hàm callback cho widget Entry
vcmd = (root.register(validate_input), "%P")

# Đặt thuộc tính validate và validatecommand cho widget Entry
entry.config(validate="key", validatecommand=vcmd)

# Gán một hàm callback cho sự kiện <Return> của widget Entry # Thêm dòng này để gọi hàm disable_entry() khi nhấn Enter
entry.bind("<Return>", lambda event: disable_entry())

root.configure(bg="lightskyblue")

# Sử dụng một vòng lặp while để lặp lại hàm quet_mv() cho đến khi biến continue_scan là False

while continue_scan:

	root.mainloop()

root.configure(bg="lightskyblue")

